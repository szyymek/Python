import openpyxl
import numpy as np
import operator

def levenshtein_ratio_and_distance(s, t, ratio_calc = False):
    """ levenshtein_ratio_and_distance:
        Calculates levenshtein distance between two strings.
        If ratio_calc = True, the function computes the
        levenshtein distance ratio of similarity between two strings
        For all i and j, distance[i,j] will contain the Levenshtein
        distance between the first i characters of s and the
        first j characters of t
    """
    # Initialize matrix of zeros
    rows = len(s)+1
    cols = len(t)+1
    distance = np.zeros((rows,cols),dtype = int)

    # Populate matrix of zeros with the indeces of each character of both strings
    for i in range(1, rows):
        for k in range(1,cols):
            distance[i][0] = i
            distance[0][k] = k

    # Iterate over the matrix to compute the cost of deletions,insertions and/or substitutions    
    for col in range(1, cols):
        for row in range(1, rows):
            if s[row-1] == t[col-1]:
                cost = 0 # If the characters are the same in the two strings in a given position [i,j] then the cost is 0
            else:
                # In order to align the results with those of the Python Levenshtein package, if we choose to calculate the ratio
                # the cost of a substitution is 2. If we calculate just distance, then the cost of a substitution is 1.
                if ratio_calc == True:
                    cost = 2
                else:
                    cost = 1
            distance[row][col] = min(distance[row-1][col] + 1,      # Cost of deletions
                                 distance[row][col-1] + 1,          # Cost of insertions
                                 distance[row-1][col-1] + cost)     # Cost of substitutions
    if ratio_calc == True:
        # Computation of the Levenshtein Distance Ratio
        Ratio = ((len(s)+len(t)) - distance[row][col]) / (len(s)+len(t))
        return Ratio
    else:
        # print(distance) # Uncomment if you want to see the matrix showing how the algorithm computes the cost of deletions,
        # insertions and/or substitutions
        # This is the minimum number of edits needed to convert string a to string b
        return "The strings are {} edits away".format(distance[row][col])
    
def max_kolumna(wiersz, plik):
    wynik=0
    for i in range(1,50):
        if plik.cell(row=wiersz,column=i).value is None:
            return wynik
        else:
            wynik+=1
  
# scieżki do plikow
path = "C:\\Users\szymon.grzybicki\Documents\Python Scripts\Indeksy wszystkie.xlsx"
path2 = "C:\\Users\szymon.grzybicki\Documents\Python Scripts\Historia ofertowania.xlsx"
path3 = "C:\\Users\szymon.grzybicki\Documents\Python Scripts\Wspolczynniki.xlsx"
path4 = "C:\\Users\szymon.grzybicki\Documents\Python Scripts\Aplikacja134.xlsx"
  
# stworzenie objektu "plik excela"
wb_obj = openpyxl.load_workbook(path)
wb_obj2 = openpyxl.load_workbook(path2)
wb_obj3 = openpyxl.load_workbook(path3)
wb_obj4 = openpyxl.load_workbook(path4)

#-------------------------definicje

indeksy = wb_obj['Opisy indeksów'] # zakładka Indeksy z pliku Dostawcy
dostawcy = wb_obj2['Hist']
podobne = wb_obj3['Opisy indeksów']
brak = wb_obj4['Arkusz1']
  
#max_col = indeksy.max_column 

producenci=[]

for j in range(57,107):
    bramka = True
    dost2 = 6
    for k in range(0, dostawcy.max_row-1):        
        ind = dostawcy.cell(row = k+2, column = 1).value
        Szukany_indeks = brak.cell(row = j+2, column = 1).value
        if ind == Szukany_indeks:
            for y in range(2,30,2):
                if dostawcy.cell(row = k+2, column = y).value is None:
                    break
                else:
                    print("Indeks ", ind, " był kupowany w: ", dostawcy.cell(row = k+2, column = y).value)
                    brak.cell(row = j+2, column = dost2).value = dostawcy.cell(row = k+2, column = y).value
                    dost2=dost2+1
            bramka = False
            break
    if bramka == False:
        continue
        
    print("Szukam rekomendacji dla: ", brak.cell(row = j+2, column = 1).value, " ", brak.cell(row = j+2, column = 2).value)
    Szukany_numer = str(brak.cell(row = j+2, column = 4).value)
    print("Szukany numer: ", Szukany_numer)
    Szukana_nazwa = str(brak.cell(row = j+2, column = 5).value)
    if Szukany_numer is None:
        continue
    if Szukana_nazwa is None:
        continue
#maska = "CS-145"

    for k in range(0, indeksy.max_row-1):
    #if indeksy.cell(row = k+2, column = 1).value[:6] == maska:
        part_number = str(indeksy.cell(row = k+2, column = 4).value)
        nazwa = str(indeksy.cell(row = k+2, column = 5).value)
        if type(part_number) is not None:
            if type(nazwa) is not None:
                podzielony_numer=part_number.split()
                indeksy.cell(row = k+2, column = 6).value = levenshtein_ratio_and_distance(part_number.lower(),Szukany_numer.lower(),ratio_calc = True)
                indeksy.cell(row = k+2, column = 7).value = levenshtein_ratio_and_distance(nazwa.lower(),Szukana_nazwa.lower(),ratio_calc = True)
                indeksy.cell(row = k+2, column = 8).value = indeksy.cell(row = k+2, column = 7).value + indeksy.cell(row = k+2, column = 6).value
    
    wb_obj.save("C:\\Users\szymon.grzybicki\Documents\Python Scripts\Wspolczynniki.xlsx")
    path3 = "C:\\Users\szymon.grzybicki\Documents\Python Scripts\Wspolczynniki.xlsx"
    wb_obj3 = openpyxl.load_workbook(path3)
    podobne = wb_obj3['Opisy indeksów']

    stats = {}
    rekomendacje = {}

    for k in range(0, podobne.max_row-1):
        stats[podobne.cell(row = k+2, column = 1).value] = podobne.cell(row = k+2, column = 8).value
    for x in range(0,25): #sprawdza najbardziej podobne indeksy
        najlepszy = max(stats.items(), key=operator.itemgetter(1))[0]
        #print(najlepszy)
        for i in range(0, dostawcy.max_row-1):
            for h in range(2,max_kolumna(i+2,dostawcy),2):
                if str(dostawcy.cell(row = i+2, column = 1).value) == najlepszy:
                    if (dostawcy.cell(row = i+2, column = h).value !="") and (dostawcy.cell(row = i+2, column = h).value in rekomendacje):
                        rekomendacje[dostawcy.cell(row = i+2, column = h).value] += dostawcy.cell(row = i+2, column = h+1).value
                    else:
                        rekomendacje[dostawcy.cell(row = i+2, column = h).value] = dostawcy.cell(row = i+2, column = h+1).value
        stats.pop(najlepszy, None)
    sorted_x = sorted(rekomendacje.items(), key=lambda kv: kv[1], reverse=True) 
    #print(sorted_x)
    dost1=6
    numerek = 1
    for m in sorted_x:
        print("nr ",numerek,": ",m[0], "siła rekomendacji: ", m[1])
        brak.cell(row = j+2, column = dost1).value = m[0]
        numerek = numerek + 1
        dost1 = dost1 +1
    path3 = "C:\\Users\szymon.grzybicki\Documents\Python Scripts\Wspolczynniki.xlsx"
    wb_obj3 = openpyxl.load_workbook(path3)
wb_obj4.save("C:\\Users\szymon.grzybicki\Documents\Python Scripts\ToJeTo3.xlsx")
