import openpyxl
from collections import Counter

path = "C:\\Users\szymon.grzybicki\Documents\Python Scripts\Indeksy wszystkie.xlsx"
path2 = "C:\\Users\szymon.grzybicki\Documents\Python Scripts\Producenci.xlsx"

wb_obj = openpyxl.load_workbook(path)
wb_obj2 = openpyxl.load_workbook(path2)

indeksy = wb_obj['Opisy indeksów']
producenci = wb_obj2['Arkusz1']

wynik=[]
produ=[]

sort_slownik={}
for i in range(2, indeksy.max_row+1):
    if indeksy.cell(row = i, column = 5).value is not None:
        z=indeksy.cell(row = i, column = 5).value.split()
        for slowo in z:
            if len(slowo)>1:
                wynik.append(slowo.lower())

def wszystkie_wyrazy(text):
    posortowane=[]

    word_counts = Counter(text)
    top = word_counts.most_common(6000)
    for s in range(0,6000):
        posortowane.append(top[s][0].lower())
        sort_slownik[top[s][0].lower()]=top[s][1]
    return posortowane

wszystko=wszystkie_wyrazy(wynik)

for i in range(2, producenci.max_row+1):
    if producenci.cell(row = i, column = 1).value is not None:
        h=producenci.cell(row = i, column = 1).value.split()
        for prodi in h:
            if len(prodi)>1:
                produ.append(prodi)
                if prodi.lower() in wszystko:
                    producenci.cell(row = i, column = 2).value=sort_slownik[prodi.lower()]
wb_obj2.save("C:\\Users\szymon.grzybicki\Documents\Python Scripts\Slowa.xlsx")
