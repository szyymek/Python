import openpyxl

path1 = "C:\\Users\szymon.grzybicki\Documents\Python Scripts\Historia ofertowania.xlsx"
path2 = "C:\\Users\szymon.grzybicki\Documents\Python Scripts\Czesci do sprawdzenia.xlsx"

wb_obj1 = openpyxl.load_workbook(path1)
wb_obj2 = openpyxl.load_workbook(path2)

dostawcy = wb_obj1['Hist']
brak = wb_obj2['Arkusz1']

for j in range(2,3418):
    dost2 = 6
    for k in range(0, dostawcy.max_row-1):        
        ind = dostawcy.cell(row = k+2, column = 1).value
        Szukany_indeks = brak.cell(row = j+2, column = 1).value
        if ind == Szukany_indeks:
            for y in range(2,30,2):
                if dostawcy.cell(row = k+2, column = y).value is None:
                    break
                else:
                    print("Indeks ", ind, " był kupowany w: ", dostawcy.cell(row = k+2, column = y).value)
                    brak.cell(row = j+2, column = dost2).value = dostawcy.cell(row = k+2, column = y).value
                    dost2=dost2+1
    wb_obj2.save("C:\\Users\szymon.grzybicki\Documents\Python Scripts\Wynik.xlsx")
